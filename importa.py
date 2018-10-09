#!/usr/bin/python3.6
# -*- coding: utf-8 -*-

#####################
#lib de excel
#####################
#pip3 install openpyxl
#pip3 install PyMySQL

import openpyxl
import pymysql
global conexao

"""
	Conecta no banco de dados
"""
conexao = pymysql.connect(host='localhost', user='root', password='root',database='banco')
cursor  = conexao.cursor()

wb = openpyxl.load_workbook('comunicador.xlsx')

for indice,grupo_contato in enumerate(wb.get_sheet_names()):
    sheet = wb.get_sheet_by_name(grupo_contato)

    print("==============================")
    print("Aba do excel: %s" % grupo_contato)
    print("==============================")

    conta_reg = 0

    for linha in range(1,sheet.max_row+1):

        nome     = sheet.cell(row=linha, column=1).value
        setor    = sheet.cell(row=linha, column=2).value
        email    = sheet.cell(row=linha, column=3).value
        telefone = sheet.cell(row=linha, column=4).value

        if nome is not None:

            if setor is not None:

                print("Linha: [%d] => Nome: %s Setor: %s Email: %s Telefone: %s" % (linha, nome, setor, email, telefone))

                """
                     Insert tabela de clientes
                """
                insert_query = ""
                insert_query = "insert into clientes (nome,setor,email,telefone) VALUES (%s,%s,%s,%s)"
                value_query = (nome,setor,email,telefone)
                cursor.execute(insert_query, value_query)
                conexao.commit()


                conta_reg = conta_reg + 1


    print("Total de Registros: %d " % conta_reg)

cursor.close()
